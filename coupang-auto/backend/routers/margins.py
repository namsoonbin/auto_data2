# -*- coding: utf-8 -*-
from fastapi import APIRouter, HTTPException, Depends, Query, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional, List
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

from models.schemas import (
    MarginCreate, MarginUpdate, MarginResponse,
    MarginListResponse, UnmatchedProductsResponse, UnmatchedProduct
)
from utils.query_helpers import escape_like_pattern
from models.auth import User, Tenant
from services.database import get_db, ProductMargin, IntegratedRecord
from auth.dependencies import get_current_user, get_current_tenant

router = APIRouter()


@router.get("/margins", response_model=MarginListResponse)
async def get_all_margins(
    limit: Optional[int] = Query(None, ge=1, le=1000, description="Limit number of records"),
    offset: Optional[int] = Query(0, ge=0, description="Offset for pagination"),
    product_name: Optional[str] = Query(None, description="Filter by product name (partial match)"),
    option_id: Optional[int] = Query(None, description="Filter by option ID"),
    min_margin_rate: Optional[float] = Query(None, description="Minimum margin rate %"),
    max_margin_rate: Optional[float] = Query(None, description="Maximum margin rate %"),
    sort_by: str = Query("updated_at", regex="^(updated_at|created_at|margin_rate|option_id|product_name)$"),
    sort_order: str = Query("desc", regex="^(asc|desc)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    current_tenant: Tenant = Depends(get_current_tenant)
):
    """
    Get all margin records with optional filtering and pagination

    Returns:
        - margins: List of margin records
        - count: Number of records returned
        - total: Total records matching filters
    """
    # Build base query - filter by tenant
    query = db.query(ProductMargin).filter(ProductMargin.tenant_id == current_tenant.id)

    # Apply filters
    if product_name:
        query = query.filter(ProductMargin.product_name.like(f"%{escape_like_pattern(product_name)}%"))

    if option_id:
        query = query.filter(ProductMargin.option_id == option_id)

    if min_margin_rate is not None:
        query = query.filter(ProductMargin.margin_rate >= min_margin_rate)

    if max_margin_rate is not None:
        query = query.filter(ProductMargin.margin_rate <= max_margin_rate)

    # Get total count before pagination
    total = query.count()

    # Apply sorting
    sort_column = getattr(ProductMargin, sort_by)
    if sort_order == "desc":
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())

    # Apply pagination
    if limit:
        query = query.offset(offset).limit(limit)

    margins = query.all()

    return MarginListResponse(
        margins=margins,
        count=len(margins),
        total=total
    )


@router.post("/margins/recalculate")
async def recalculate_margin_data(
    start_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    current_tenant: Tenant = Depends(get_current_tenant)
):
    """
    날짜 범위를 선택하여 IntegratedRecord의 마진 데이터를 현재 ProductMargin 데이터로 재계산

    Parameters:
        - start_date: 시작 날짜 (선택사항, 미지정시 전체 기간)
        - end_date: 종료 날짜 (선택사항, 미지정시 전체 기간)

    Returns:
        - updated_count: 업데이트된 레코드 수
        - matched_products: 마진 데이터와 매칭된 상품 수
        - total_records: 대상 레코드 총 수
    """
    try:
        # Parse dates if provided
        date_filter_start = None
        date_filter_end = None

        if start_date:
            try:
                date_filter_start = datetime.strptime(start_date, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail="시작 날짜 형식이 올바르지 않습니다. YYYY-MM-DD 형식을 사용하세요"
                )

        if end_date:
            try:
                date_filter_end = datetime.strptime(end_date, "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail="종료 날짜 형식이 올바르지 않습니다. YYYY-MM-DD 형식을 사용하세요"
                )

        # Build query for IntegratedRecord - filter by tenant
        query = db.query(IntegratedRecord).filter(IntegratedRecord.tenant_id == current_tenant.id)

        if date_filter_start:
            query = query.filter(IntegratedRecord.date >= date_filter_start)
        if date_filter_end:
            query = query.filter(IntegratedRecord.date <= date_filter_end)

        # Get all records in date range
        records = query.all()
        total_records = len(records)

        if total_records == 0:
            return {
                "status": "success",
                "message": "선택한 기간에 레코드가 없습니다",
                "total_records": 0,
                "updated_count": 0,
                "matched_products": 0
            }

        # Get all margin data for this tenant
        margins = db.query(ProductMargin).filter(ProductMargin.tenant_id == current_tenant.id).all()
        margin_dict = {m.option_id: m for m in margins}

        if not margin_dict:
            return {
                "status": "warning",
                "message": "데이터베이스에 마진 데이터가 없습니다. 먼저 마진 데이터를 추가해주세요.",
                "total_records": total_records,
                "updated_count": 0,
                "matched_products": 0
            }

        # Update records
        updated_count = 0
        matched_products = set()

        for record in records:
            margin = margin_dict.get(record.option_id)

            if margin:
                # Update margin data
                record.cost_price = margin.cost_price
                record.selling_price = margin.selling_price
                record.margin_amount = margin.margin_amount
                record.margin_rate = margin.margin_rate
                record.fee_rate = margin.fee_rate
                record.fee_amount = margin.fee_amount
                record.vat = margin.vat

                # Recalculate metrics
                record.calculate_metrics()
                record.updated_at = datetime.now()

                updated_count += 1
                matched_products.add(record.option_id)

        # Commit changes
        db.commit()

        return {
            "status": "success",
            "message": f"{updated_count}개 레코드를 성공적으로 재계산했습니다 (고유 상품 {len(matched_products)}개)",
            "total_records": total_records,
            "updated_count": updated_count,
            "matched_products": len(matched_products),
            "date_range": {
                "start": start_date or "전체",
                "end": end_date or "전체"
            }
        }

    except Exception as e:
        db.rollback()
        import traceback
        error_detail = traceback.format_exc()
        print(f"Recalculation error: {error_detail}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to recalculate margin data: {str(e)}"
        )


@router.get("/margins/{option_id}", response_model=MarginResponse)
async def get_margin_by_option_id(
    option_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    current_tenant: Tenant = Depends(get_current_tenant)
):
    """Get specific margin record by option_id"""
    margin = db.query(ProductMargin).filter(
        ProductMargin.option_id == option_id,
        ProductMargin.tenant_id == current_tenant.id
    ).first()

    if not margin:
        raise HTTPException(
            status_code=404,
            detail=f"Margin record not found for option_id: {option_id}"
        )

    return margin


@router.post("/margins", response_model=MarginResponse, status_code=201)
async def create_margin(
    margin: MarginCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    current_tenant: Tenant = Depends(get_current_tenant)
):
    """
    Create new margin record

    Validates:
        - option_id is unique
        - product_name is not empty
        - numeric fields are non-negative
    """
    # Check if option_id already exists for this tenant
    existing = db.query(ProductMargin).filter(
        ProductMargin.option_id == margin.option_id,
        ProductMargin.tenant_id == current_tenant.id
    ).first()

    if existing:
        raise HTTPException(
            status_code=400,
            detail=f"Margin record already exists for option_id: {margin.option_id}"
        )

    try:
        # Create new record with tenant_id
        new_margin = ProductMargin(
            tenant_id=current_tenant.id,
            option_id=margin.option_id,
            product_name=margin.product_name,
            option_name=margin.option_name,
            cost_price=margin.cost_price,
            selling_price=margin.selling_price,
            margin_rate=margin.margin_rate,
            fee_rate=margin.fee_rate,
            fee_amount=margin.fee_amount,
            vat=margin.vat,
            notes=margin.notes
        )

        db.add(new_margin)
        db.commit()
        db.refresh(new_margin)

        return new_margin

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to create margin record: {str(e)}"
        )


@router.put("/margins/{option_id}", response_model=MarginResponse)
async def update_margin(
    option_id: int,
    margin_update: MarginUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    current_tenant: Tenant = Depends(get_current_tenant)
):
    """
    Update existing margin record

    Only provided fields will be updated (partial update)
    """
    # Find existing record for this tenant
    margin = db.query(ProductMargin).filter(
        ProductMargin.option_id == option_id,
        ProductMargin.tenant_id == current_tenant.id
    ).first()

    if not margin:
        raise HTTPException(
            status_code=404,
            detail=f"Margin record not found for option_id: {option_id}"
        )

    try:
        # Update only provided fields
        update_data = margin_update.model_dump(exclude_unset=True)

        for field, value in update_data.items():
            setattr(margin, field, value)

        # Update timestamp
        margin.updated_at = datetime.now()

        db.commit()
        db.refresh(margin)

        return margin

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to update margin record: {str(e)}"
        )


@router.delete("/margins/{option_id}")
async def delete_margin(
    option_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    current_tenant: Tenant = Depends(get_current_tenant)
):
    """Delete specific margin record by option_id"""
    margin = db.query(ProductMargin).filter(
        ProductMargin.option_id == option_id,
        ProductMargin.tenant_id == current_tenant.id
    ).first()

    if not margin:
        raise HTTPException(
            status_code=404,
            detail=f"Margin record not found for option_id: {option_id}"
        )

    try:
        db.delete(margin)
        db.commit()

        return {
            "message": f"Margin record deleted successfully for option_id: {option_id}",
            "option_id": option_id
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to delete margin record: {str(e)}"
        )


@router.delete("/margins")
async def delete_all_margins(
    confirm: bool = Query(False, description="Must be true to confirm deletion"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    current_tenant: Tenant = Depends(get_current_tenant)
):
    """
    Delete all margin records (requires confirmation)

    Safety feature: requires confirm=true query parameter
    """
    if not confirm:
        raise HTTPException(
            status_code=400,
            detail="Deletion not confirmed. Add ?confirm=true to delete all records"
        )

    try:
        deleted_count = db.query(ProductMargin).filter(
            ProductMargin.tenant_id == current_tenant.id
        ).delete()
        db.commit()

        return {
            "message": f"Successfully deleted {deleted_count} margin records",
            "deleted_count": deleted_count
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to delete margin records: {str(e)}"
        )


@router.get("/margins/unmatched/products", response_model=UnmatchedProductsResponse)
async def get_unmatched_products(
    min_sales: Optional[float] = Query(0, ge=0, description="Minimum sales amount filter"),
    limit: Optional[int] = Query(100, ge=1, le=1000, description="Limit results"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    current_tenant: Tenant = Depends(get_current_tenant)
):
    """
    Find products in IntegratedRecord that don't have margin data

    Useful for identifying products that need margin configuration

    Returns products sorted by sales_amount (highest first)
    """
    # Get all option_ids that have margin data for this tenant
    margin_option_ids = db.query(ProductMargin.option_id).filter(
        ProductMargin.tenant_id == current_tenant.id
    ).all()
    margin_option_ids = [row[0] for row in margin_option_ids]

    # Query IntegratedRecord for products without margin data for this tenant
    query = db.query(
        IntegratedRecord.option_id,
        IntegratedRecord.product_name,
        IntegratedRecord.option_name,
        IntegratedRecord.sales_amount,
        IntegratedRecord.sales_quantity
    ).filter(
        IntegratedRecord.tenant_id == current_tenant.id,
        IntegratedRecord.option_id.notin_(margin_option_ids) if margin_option_ids else True,
        IntegratedRecord.sales_amount >= min_sales
    ).order_by(
        IntegratedRecord.sales_amount.desc()
    )

    if limit:
        query = query.limit(limit)

    results = query.all()

    # Convert to UnmatchedProduct objects
    unmatched = [
        UnmatchedProduct(
            option_id=row.option_id,
            product_name=row.product_name,
            option_name=row.option_name,
            sales_amount=row.sales_amount,
            sales_quantity=row.sales_quantity
        )
        for row in results
    ]

    return UnmatchedProductsResponse(
        unmatched_products=unmatched,
        count=len(unmatched)
    )


@router.get("/margins/template/download")
async def download_margin_template(
    current_user: User = Depends(get_current_user),
    current_tenant: Tenant = Depends(get_current_tenant)
):
    """
    Download Excel template for margin data bulk upload

    Returns an Excel file with proper column headers and sample data
    """
    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "마진 데이터 템플릿"

    # Define headers (Korean + English descriptions)
    headers = [
        ("option_id", "옵션 ID"),
        ("product_name", "상품명"),
        ("option_name", "옵션명"),
        ("cost_price", "도매가(원가)"),
        ("selling_price", "판매가"),
        ("margin_rate", "마진율(%)"),
        ("fee_rate", "수수료율(%)"),
        ("fee_amount", "수수료액"),
        ("vat", "부가세"),
        ("notes", "비고")
    ]

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers (Row 1: English, Row 2: Korean)
    for col_idx, (eng, kor) in enumerate(headers, 1):
        # English header
        cell = ws.cell(row=1, column=col_idx, value=eng)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

        # Korean description in row 2
        cell2 = ws.cell(row=2, column=col_idx, value=kor)
        cell2.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell2.font = Font(bold=True, size=10)
        cell2.alignment = header_alignment
        cell2.border = border

    # Add sample data (row 3)
    sample_data = [
        123456789,  # option_id
        "샘플 상품명",  # product_name
        "샘플 옵션명",  # option_name
        10000,  # cost_price
        20000,  # selling_price
        40.0,  # margin_rate
        10.0,  # fee_rate
        2000,  # fee_amount
        1000,  # vat
        "샘플 데이터"  # notes
    ]

    for col_idx, value in enumerate(sample_data, 1):
        cell = ws.cell(row=3, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border

    # Adjust column widths
    column_widths = {
        'A': 15,  # option_id
        'B': 30,  # product_name
        'C': 25,  # option_name
        'D': 15,  # cost_price
        'E': 15,  # selling_price
        'F': 12,  # margin_rate
        'G': 15,  # fee_rate
        'H': 12,  # fee_amount
        'I': 12,  # vat
        'J': 30   # notes
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header rows
    ws.freeze_panes = "A3"

    # Add instruction sheet
    ws_info = wb.create_sheet("사용 방법")
    instructions = [
        ["마진 데이터 Excel 템플릿 사용 방법"],
        [""],
        ["1. 필수 컬럼 (반드시 입력해야 함):"],
        ["   - option_id: 옵션 ID (숫자, 필수)"],
        ["   - product_name: 상품명 (텍스트, 필수)"],
        ["   - cost_price: 도매가/원가 (숫자, 필수, 0보다 커야 함)"],
        ["   - fee_amount: 총 수수료액 (숫자, 필수, 0 이상)"],
        ["   - vat: 부가세 (숫자, 필수, 0 이상)"],
        [""],
        ["   ⚠️ 위 항목들은 순이익 계산에 필수입니다!"],
        ["   ⚠️ 입력하지 않으면 업로드가 거부됩니다."],
        [""],
        ["2. 선택 컬럼:"],
        ["   - option_name: 옵션명"],
        ["   - selling_price: 판매가"],
        ["   - margin_rate: 마진율(%)"],
        ["   - fee_rate: 수수료율(%)"],
        ["   - notes: 비고"],
        [""],
        ["3. 백분율 입력 방법:"],
        ["   - margin_rate, fee_rate는 백분율(%) 또는 소수로 입력 가능"],
        ["   - 백분율 형식: 10%, 20%, 30% (셀 서식을 백분율로 설정)"],
        ["   - 소수 형식: 10, 20, 30 (일반 숫자로 입력)"],
        ["   - 시스템이 자동으로 형식을 감지하여 변환합니다"],
        [""],
        ["4. 주의사항:"],
        ["   - 첫 번째 행(영문 컬럼명)과 두 번째 행(한글 설명)은 삭제하지 마세요"],
        ["   - 세 번째 행의 샘플 데이터는 삭제하고 실제 데이터를 입력하세요"],
        ["   - option_id는 중복되지 않아야 합니다"],
        ["   - 숫자 필드는 숫자만 입력하세요"],
        [""],
        ["5. 업로드 방법:"],
        ["   - 마진 관리 페이지에서 'Excel 업로드' 버튼 클릭"],
        ["   - 작성한 파일 선택 후 업로드"]
    ]

    for row_idx, row_data in enumerate(instructions, 1):
        ws_info.cell(row=row_idx, column=1, value=row_data[0])

    ws_info.column_dimensions['A'].width = 80

    # Title formatting
    ws_info.cell(row=1, column=1).font = Font(bold=True, size=14, color="4472C4")

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Return as downloadable file
    filename = f"margin_template_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/margins/upload/excel")
async def upload_margin_excel(
    file: UploadFile = File(..., description="Margin data Excel file"),
    skip_existing: bool = Query(True, description="Skip records with existing option_id"),
    update_existing: bool = Query(False, description="Update existing records instead of skipping"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    current_tenant: Tenant = Depends(get_current_tenant)
):
    """
    Upload Excel file with margin data and bulk create/update records

    Excel file should have columns:
    - option_id (required)
    - product_name (required)
    - option_name, cost_price, selling_price, margin_rate,
      fee_rate, fee_amount, vat, notes (optional)

    Returns:
        - created: Number of new records created
        - updated: Number of existing records updated
        - skipped: Number of records skipped
        - errors: List of errors encountered
    """
    # Validate file format
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Invalid file format. Only .xlsx and .xls files are supported"
        )

    try:
        # Read Excel file
        df = pd.read_excel(file.file)

        # Check for required columns (including critical margin data)
        required_columns = ['option_id', 'product_name', 'cost_price', 'fee_amount', 'vat']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {', '.join(missing_columns)}. These fields are mandatory for accurate profit calculation."
            )

        # Optional columns with defaults
        optional_columns = {
            'option_name': '',
            'cost_price': 0.0,
            'selling_price': 0.0,
            'margin_rate': 0.0,
            'fee_rate': 0.0,
            'fee_amount': 0.0,
            'vat': 0.0,
            'notes': ''
        }

        # Fill missing optional columns
        for col, default in optional_columns.items():
            if col not in df.columns:
                df[col] = default

        # Remove rows where required fields are NaN
        df = df.dropna(subset=['option_id', 'product_name', 'cost_price', 'fee_amount', 'vat'])

        # Convert option_id to int
        df['option_id'] = pd.to_numeric(df['option_id'], errors='coerce')
        df = df.dropna(subset=['option_id'])
        df['option_id'] = df['option_id'].astype('int64')

        # Validate required numeric fields are greater than 0
        df['cost_price'] = pd.to_numeric(df['cost_price'], errors='coerce')
        df['fee_amount'] = pd.to_numeric(df['fee_amount'], errors='coerce')
        df['vat'] = pd.to_numeric(df['vat'], errors='coerce')

        # Check for invalid values in required fields (must be > 0 for cost_price, >= 0 for others)
        invalid_cost = df[df['cost_price'] <= 0]
        invalid_fee = df[df['fee_amount'] < 0]
        invalid_vat = df[df['vat'] < 0]

        validation_errors = []
        if len(invalid_cost) > 0:
            validation_errors.append(f"{len(invalid_cost)} rows have invalid cost_price (must be > 0)")
        if len(invalid_fee) > 0:
            validation_errors.append(f"{len(invalid_fee)} rows have negative fee_amount")
        if len(invalid_vat) > 0:
            validation_errors.append(f"{len(invalid_vat)} rows have negative vat")

        if validation_errors:
            raise HTTPException(
                status_code=400,
                detail=f"Validation failed: {'; '.join(validation_errors)}"
            )

        # Remove any rows with NaN in required numeric fields after conversion
        df = df.dropna(subset=['cost_price', 'fee_amount', 'vat'])

        # Fill NaN values in numeric columns
        numeric_cols = ['cost_price', 'selling_price', 'margin_rate',
                       'fee_rate', 'fee_amount', 'vat']
        for col in numeric_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)

        # Convert percentage columns from Excel format (0.1 → 10%)
        # Excel stores percentages as decimals (10% = 0.1)
        percentage_cols = ['margin_rate', 'fee_rate']
        for col in percentage_cols:
            if col in df.columns:
                # Check if values are in decimal format (likely from Excel percentage format)
                # If most values are < 1, assume they're Excel percentages and multiply by 100
                sample_values = df[col][df[col] > 0]
                if len(sample_values) > 0 and sample_values.median() < 1:
                    df[col] = df[col] * 100
                    print(f"Converted {col} from Excel percentage format (multiplied by 100)")

        # Fill NaN values in string columns
        string_cols = ['option_name', 'product_name', 'notes']
        for col in string_cols:
            df[col] = df[col].fillna('').astype(str)

        print(f"Parsed {len(df)} records from Excel file")

        # Process records
        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                option_id = int(row['option_id'])

                # Check if record exists for this tenant
                existing = db.query(ProductMargin).filter(
                    ProductMargin.option_id == option_id,
                    ProductMargin.tenant_id == current_tenant.id
                ).first()

                if existing:
                    if update_existing:
                        # Update existing record
                        existing.product_name = str(row['product_name'])
                        existing.option_name = str(row['option_name'])
                        existing.cost_price = float(row['cost_price'])
                        existing.selling_price = float(row['selling_price'])
                        existing.margin_rate = float(row['margin_rate'])
                        existing.fee_rate = float(row['fee_rate'])
                        existing.fee_amount = float(row['fee_amount'])
                        existing.vat = float(row['vat'])
                        existing.notes = str(row['notes'])
                        existing.updated_at = datetime.now()
                        updated_count += 1
                    elif skip_existing:
                        skipped_count += 1
                        continue
                    else:
                        errors.append(f"Row {idx + 3}: option_id {option_id} already exists")
                        continue
                else:
                    # Create new record with tenant_id
                    new_margin = ProductMargin(
                        tenant_id=current_tenant.id,
                        option_id=option_id,
                        product_name=str(row['product_name']),
                        option_name=str(row['option_name']),
                        cost_price=float(row['cost_price']),
                        selling_price=float(row['selling_price']),
                        margin_rate=float(row['margin_rate']),
                        fee_rate=float(row['fee_rate']),
                        fee_amount=float(row['fee_amount']),
                        vat=float(row['vat']),
                        notes=str(row['notes'])
                    )
                    db.add(new_margin)
                    created_count += 1

            except Exception as e:
                errors.append(f"Row {idx + 3}: {str(e)}")
                continue

        # Commit all changes
        try:
            db.commit()
            print(f"Successfully committed: {created_count} created, {updated_count} updated")
        except Exception as e:
            db.rollback()
            raise HTTPException(
                status_code=500,
                detail=f"Failed to commit changes to database: {str(e)}"
            )

        # Determine status
        if len(errors) == 0:
            status = "success"
            message = f"Successfully processed {created_count + updated_count} records"
        elif created_count + updated_count > 0:
            status = "partial_success"
            message = f"Partially successful: {created_count} created, {updated_count} updated, {len(errors)} errors"
        else:
            status = "error"
            message = f"Failed to process records: {len(errors)} errors"

        return {
            "status": status,
            "message": message,
            "created": created_count,
            "updated": updated_count,
            "skipped": skipped_count,
            "total_processed": created_count + updated_count,
            "errors": errors[:10] if len(errors) > 10 else errors,  # Limit to first 10 errors
            "total_errors": len(errors)
        }

    except pd.errors.ParserError as e:
        raise HTTPException(
            status_code=400,
            detail=f"Failed to parse Excel file: {str(e)}"
        )
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"Excel upload error: {error_detail}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to process Excel file: {str(e)}"
        )


@router.post("/margins/batch", response_model=dict)
async def batch_create_margins(
    margins: List[MarginCreate],
    skip_existing: bool = Query(True, description="Skip records with existing option_id"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    current_tenant: Tenant = Depends(get_current_tenant)
):
    """
    Batch create multiple margin records

    Parameters:
        - margins: List of margin records to create
        - skip_existing: If true, skip existing option_ids; if false, raise error

    Returns:
        - created: Number of records created
        - skipped: Number of records skipped
        - errors: List of errors encountered
    """
    created_count = 0
    skipped_count = 0
    errors = []

    for margin in margins:
        try:
            # Check if exists for this tenant
            existing = db.query(ProductMargin).filter(
                ProductMargin.option_id == margin.option_id,
                ProductMargin.tenant_id == current_tenant.id
            ).first()

            if existing:
                if skip_existing:
                    skipped_count += 1
                    continue
                else:
                    errors.append(f"option_id {margin.option_id} already exists")
                    continue

            # Create new record with tenant_id
            new_margin = ProductMargin(
                tenant_id=current_tenant.id,
                option_id=margin.option_id,
                product_name=margin.product_name,
                option_name=margin.option_name,
                cost_price=margin.cost_price,
                selling_price=margin.selling_price,
                margin_rate=margin.margin_rate,
                fee_rate=margin.fee_rate,
                fee_amount=margin.fee_amount,
                vat=margin.vat,
                notes=margin.notes
            )

            db.add(new_margin)
            created_count += 1

        except Exception as e:
            errors.append(f"option_id {margin.option_id}: {str(e)}")
            continue

    # Commit all changes
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to commit batch create: {str(e)}"
        )

    return {
        "status": "success" if len(errors) == 0 else "partial_success",
        "message": f"Created {created_count} records, skipped {skipped_count}, {len(errors)} errors",
        "created": created_count,
        "skipped": skipped_count,
        "errors": errors
    }
